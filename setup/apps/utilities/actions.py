from django.utils.translation import gettext_lazy as _
from django.utils.timezone import datetime
from django.http import HttpResponse
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from math import ceil


def export_to_excel_file(admin_class, request, queryset):
    excel_file = Workbook()
    sheet_1 = excel_file.active
    sheet_1.title = _("Main").encode("utf8")
    columns_size = []

    headers = queryset.model._meta.fields
    for i, header in enumerate(headers, start=1):
        sheet_1[f'{get_column_letter(i)}1'] = header.verbose_name.encode(
            "utf8")
        columns_size.append(15)

    for i, row in enumerate(queryset, start=1):
        for j, field in enumerate(headers):
            cell_coordinates = f'{get_column_letter(j + 1)}{i + 1}'
            data = getattr(row, field.name)
            display = getattr(row, f"get_{field.name}_display", None)
            if display:
                data = display()
            elif type(data) == datetime:
                data = data.strftime("%Y-%m-%d %H:%M:%S")
            else:
                data = str(data)

            sheet_1[cell_coordinates] = data.encode("utf8") if j != 0 else i

            if columns_size[j] < ceil(len(data) * 1.5):
                columns_size[j] = ceil(len(data) * 1.5)

    for i, column_size in enumerate(columns_size, start=1):
        sheet_1.column_dimensions[get_column_letter(i)].width = column_size

    excel_bytes = save_virtual_workbook(excel_file)

    filename = 'my-excel-%s.xlsx' % datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(excel_bytes, content_type=content_type)
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return response


class ExcelAdminMixin:

    def get_actions(self, request):
        actions = super().get_actions(request)

        action_name = "export_to_excel_file"
        actions[action_name] = (
            export_to_excel_file,
            action_name,
            _("Export to excel file"),)
        return actions
